import logging
import os
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generator for Excel files from spreadsheet data."""

    def __init__(self, temp_dir: str = '/tmp'):
        """
        Initialize Excel Generator.

        Args:
            temp_dir: Directory for temporary Excel files
        """
        self.temp_dir = temp_dir
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir, exist_ok=True)
            logger.info(f"Created temp directory: {temp_dir}")

    async def create_excel(self, data: List[Dict[str, str]]) -> Optional[str]:
        """
        Create Excel file from data.

        Args:
            data: List of dictionaries with client data

        Returns:
            Path to created Excel file or None
        """
        try:
            if not data:
                logger.warning("No data provided for Excel generation")
                return None

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Клиенты"

            headers = ['ID', 'Дата регистрации', 'Email', 'Телефон', 'Процент', 'Роль', 'Агент']
            worksheet.append(headers)

            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            header_font = Font(
                bold=True,
                color="FFFFFF",
                size=12
            )

            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row_num, row_data in enumerate(data, start=2):
                date_val = row_data.get('Дата', '')
                phone_val = row_data.get('Телефон', '')
                email_val = row_data.get('Email', '')
                percentage_val = row_data.get('Процент', '')

                worksheet.cell(row=row_num, column=1, value=date_val)
                worksheet.cell(row=row_num, column=2, value=phone_val)
                worksheet.cell(row=row_num, column=3, value=email_val)
                worksheet.cell(row=row_num, column=4, value=percentage_val)

                for col_num in range(1, 5):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True
                    )

            for header_cell in worksheet[1]:
                header_cell.border = thin_border

            worksheet.column_dimensions['A'].width = 8
            worksheet.column_dimensions['B'].width = 18
            worksheet.column_dimensions['C'].width = 25
            worksheet.column_dimensions['D'].width = 18
            worksheet.column_dimensions['E'].width = 12
            worksheet.column_dimensions['F'].width = 12
            worksheet.column_dimensions['G'].width = 25

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"clients_{timestamp}.xlsx"
            filepath = os.path.join(self.temp_dir, filename)

            workbook.save(filepath)

            logger.info(f"Excel file created: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return None

    def cleanup_temp_files(self, max_age_hours: int = 1) -> None:
        """
        Clean up old temporary Excel files.

        Args:
            max_age_hours: Maximum age of files to keep in hours
        """
        try:
            import glob
            from datetime import timedelta

            cutoff_time = datetime.now() - timedelta(hours=max_age_hours)

            for filepath in glob.glob(os.path.join(self.temp_dir, 'clients_*.xlsx')):
                try:
                    file_time = datetime.fromtimestamp(os.path.getmtime(filepath))
                    if file_time < cutoff_time:
                        os.remove(filepath)
                        logger.info(f"Cleaned up old file: {filepath}")
                except Exception as e:
                    logger.warning(f"Error removing file {filepath}: {e}")

        except Exception as e:
            logger.error(f"Error cleaning up temp files: {e}")
